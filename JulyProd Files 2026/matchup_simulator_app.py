#!/usr/bin/env python3
"""
Matchup Simulator Web Application
Flask-based UI for running matchmaking simulations with detailed reporting
"""

import json
import os
import sys
from collections import defaultdict, Counter
from datetime import datetime

from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Setup paths
GAME_DIR = r"C:\BPClone_Claude"
LEAGUE_BASE_DIR = r"C:\BPClone_Claude\saves\league"

sys.path.insert(0, GAME_DIR)
os.chdir(GAME_DIR)

from team import Team
from matchmaking import build_global_fight_card
from warrior import get_warrior_tier, TIER_ORDER
from file_protection import load_json_protected

app = Flask(__name__)

# Global simulation results storage
simulation_results = {}
AVAILABLE_TURNS = []  # Will be populated after function definition

def get_available_turns():
    """Get list of available turn numbers."""
    try:
        dirs = [d for d in os.listdir(LEAGUE_BASE_DIR)
                if os.path.isdir(os.path.join(LEAGUE_BASE_DIR, d)) and d.startswith('turn_')]
        turns = sorted([int(d.split('_')[1]) for d in dirs])
        return turns
    except Exception as e:
        print(f"Error getting turns: {e}")
        return []

def load_uploaded_teams(turn_num):
    """Load all uploaded teams from specified turn directory."""
    teams = []
    upload_dir = os.path.join(LEAGUE_BASE_DIR, f"turn_{turn_num:04d}")

    if not os.path.exists(upload_dir):
        return teams, f"Turn {turn_num} directory not found"

    try:
        files = [f for f in os.listdir(upload_dir) if f.startswith('upload_') and f.endswith('.json')]
    except Exception as e:
        return teams, str(e)

    for filename in sorted(files):
        try:
            team_file = os.path.join(upload_dir, filename)
            upload_data = load_json_protected(team_file)
            team_data = upload_data.get('team', upload_data)
            team = Team.from_dict(team_data)
            teams.append(team)
        except Exception as e:
            print(f"[ERROR] Failed to load {filename}: {e}")

    return teams, None

def calculate_warrior_record(warrior):
    """Get warrior's win-loss record."""
    wins = getattr(warrior, 'wins', 0)
    losses = getattr(warrior, 'losses', 0)
    return f"{wins}-{losses}"

def run_simulation(turn_num):
    """Run the matchmaking simulation for a specific turn."""
    teams, error = load_uploaded_teams(turn_num)
    if error:
        return None, f"Error loading teams: {error}"

    if not teams:
        return None, "No teams loaded"

    try:
        fights = build_global_fight_card(teams, teams, champion_state={})
        return fights, None
    except Exception as e:
        return None, f"Simulation error: {str(e)}"

def organize_fights(fights):
    """Organize fights by type."""
    fights_by_type = defaultdict(list)
    for fight in fights:
        fights_by_type[fight.fight_type].append(fight)
    return dict(fights_by_type)

# Tiers ordered lowest -> highest (TIER_ORDER itself is highest -> lowest, index 0 = Champion)
TIER_LOW_TO_HIGH = list(reversed(TIER_ORDER))

# Only fight types where both sides have a meaningful competitive tier and the
# tier-adjacency rule (challenge_tier_allowed: same tier or one tier up) actually
# applies. Peasant/Monster opponents don't have a real tier - they're fallback
# fodder, not part of the tier ladder - so they're excluded from this breakdown
# entirely (they still show in the fight-type stat cards).
TIER_BAND_FIGHT_TYPES = {"standard", "challenge", "blood_challenge"}

# Fight types whose experience gap is a fair reflection of matchmaking quality.
# Blood challenges are excluded from gap stats (though still shown in the tier
# listing itself) - they're vengeance-driven, matching a specific killer with
# zero tier/experience constraint, so folding them into an "average gap" would
# unfairly conflate an unconstrained matchup with the tier-adjacency-bound ones.
GAP_ANALYSIS_FIGHT_TYPES = {"standard", "challenge"}

def organize_fights_by_tier(fights):
    """
    Break competitive fights (standard/challenge/blood_challenge) into ordered
    sections: same-tier matchups for each tier from Recruits up to Champion,
    with the cross-tier matchups against the next tier up inserted between them
    - e.g. Recruits, Recruits/Rookies, Rookies, Rookies/Initiates, ...

    Returns a list of {"label": str, "fights": [ScheduledFight, ...]} dicts in
    display order. Any fight whose tier pair isn't same-tier or adjacent (e.g. a
    champion challenge from a non-Elites tier, since eligible challenger tiers
    are dynamic) is collected into a trailing "Other / Non-Adjacent" section so
    nothing is silently dropped.
    """
    competitive = [f for f in fights if f.fight_type in TIER_BAND_FIGHT_TYPES]

    tier_cache = {}
    def tiers_of(fight):
        key = id(fight)
        if key not in tier_cache:
            tier_cache[key] = (get_warrior_tier(fight.player_warrior), get_warrior_tier(fight.opponent))
        return tier_cache[key]

    sections = []
    used = set()

    for i, tier in enumerate(TIER_LOW_TO_HIGH):
        same = [f for f in competitive if tiers_of(f) == (tier, tier)]
        if same:
            sections.append({"label": f"{tier} Matchups", "fights": same})
            used.update(id(f) for f in same)

        if i + 1 < len(TIER_LOW_TO_HIGH):
            next_tier = TIER_LOW_TO_HIGH[i + 1]
            cross = [f for f in competitive
                     if set(tiers_of(f)) == {tier, next_tier}]
            if cross:
                sections.append({"label": f"{tier} / {next_tier} Matchups", "fights": cross})
                used.update(id(f) for f in cross)

    leftover = [f for f in competitive if id(f) not in used]
    if leftover:
        sections.append({"label": "Other / Non-Adjacent Matchups", "fights": leftover})

    # Attach experience-gap stats to each section (used by the "Experience Gap
    # by Tier" meter view and the per-fight Gap column), computed only from
    # standard/challenge fights - see GAP_ANALYSIS_FIGHT_TYPES. Blood challenge
    # fights still appear in the section's fight listing, just not counted
    # toward the average. If a section ends up with no gap-relevant fights
    # (e.g. an all-blood-challenge section), it's left without gap stats
    # entirely - same as the flat Peasant/Monster sections.
    for section in sections:
        gap_fights = [f for f in section["fights"] if f.fight_type in GAP_ANALYSIS_FIGHT_TYPES]
        if gap_fights:
            gaps = [_experience_gap(f) for f in gap_fights]
            section["avg_gap"] = sum(gaps) / len(gaps)
            section["max_gap"] = max(gaps)
            section["gap_status"] = _gap_status(section["avg_gap"])

    return sections

def _experience_gap(fight) -> int:
    """Absolute difference in total_fights between the two warriors in a fight."""
    return abs(fight.player_warrior.total_fights - fight.opponent.total_fights)

def build_flat_section(fights, fight_type, label):
    """
    Build a single, non-tier-banded section for one fight_type (Peasant/Monster
    fights) - these don't have a meaningful competitive tier so they're excluded
    from the tier ladder in organize_fights_by_tier(), but they still need to be
    listed somewhere so they aren't silently dropped from the report. Returns
    None if there are no fights of this type this turn.

    No gap stats are attached here (unlike organize_fights_by_tier's sections):
    Peasant/Monster opponents are freshly generated each fight and never
    accumulate a real total_fights count (it stays at 0), so an "experience
    gap" against them would just restate the player warrior's own fight count,
    not a genuine matchup imbalance.
    """
    matching = [f for f in fights if f.fight_type == fight_type]
    if not matching:
        return None
    return {"label": label, "fights": matching}

# Status thresholds for average experience gap (in fights). Matches the
# validated status palette: good/warning/serious/critical, each paired with an
# icon + label since color alone isn't a reliable signal (warning/serious don't
# clear 3:1 contrast on a light surface).
GAP_STATUS_LEVELS = [
    (1.5, "good",     "#0ca30c", "✓ Good"),
    (3.0, "warning",  "#fab219", "~ Fair"),
    (5.0, "serious",  "#ec835a", "! High"),
    (float("inf"), "critical", "#d03b3b", "✕ Poor"),
]

def _gap_status(avg_gap: float) -> dict:
    """Map an average experience gap to a status level (color + icon label)."""
    for threshold, key, color, label in GAP_STATUS_LEVELS:
        if avg_gap <= threshold:
            return {"key": key, "color": color, "label": label}
    return {"key": "critical", "color": GAP_STATUS_LEVELS[-1][2], "label": GAP_STATUS_LEVELS[-1][3]}

def calculate_manager_matchups(fights):
    """Calculate manager vs manager matchup counts."""
    manager_pairs = Counter()

    for fight in fights:
        if fight.opponent_manager == "The Arena":
            continue  # Skip peasant fights

        m1 = fight.player_team.manager_name
        m2 = fight.opponent_manager

        # Create a sorted tuple for consistent counting
        pair = tuple(sorted([m1, m2]))
        manager_pairs[pair] += 1

    return dict(sorted(manager_pairs.items(), key=lambda x: x[1], reverse=True))

def find_repeated_matchups(fights):
    """Find warrior pairs that match multiple times."""
    warrior_pairs = Counter()

    for fight in fights:
        w1 = fight.player_warrior.name
        w2 = fight.opponent.name

        pair = tuple(sorted([w1, w2]))
        warrior_pairs[pair] += 1

    # Return only pairs that match more than once
    repeated = {pair: count for pair, count in warrior_pairs.items() if count > 1}
    return dict(sorted(repeated.items(), key=lambda x: x[1], reverse=True))

def create_excel_export(turn, fights, organized_fights, manager_matchups, repeated_matchups, all_sections):
    """Create Excel file with simulation results."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    tier_colors = {
        "CHAMPION": "FF0000",
        "ELITES": "FF6600",
        "EXPERTS": "FFCC00",
        "VETERANS": "00CC00",
        "ADEPTS": "0066FF",
        "INITIATES": "6600FF",
        "ROOKIES": "999999",
        "RECRUITS": "CCCCCC"
    }
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== SUMMARY SHEET =====
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = f"MATCHUP SIMULATION SUMMARY - TURN {turn}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    ws_summary['A4'] = "Fight Statistics"
    ws_summary['A4'].font = Font(bold=True, size=12)

    row = 5
    ws_summary[f'A{row}'] = "Fight Type"
    ws_summary[f'B{row}'] = "Count"
    for col in ['A', 'B']:
        ws_summary[f'{col}{row}'].fill = header_fill
        ws_summary[f'{col}{row}'].font = header_font

    row += 1
    for ftype in ["blood_challenge", "challenge", "monster", "standard", "peasant"]:
        if ftype in organized_fights:
            ws_summary[f'A{row}'] = ftype.upper()
            ws_summary[f'B{row}'] = len(organized_fights[ftype])
            row += 1

    ws_summary[f'A{row}'] = "TOTAL"
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = len(fights)
    ws_summary[f'B{row}'].font = Font(bold=True)

    # ===== MANAGER VS MANAGER SHEET =====
    ws_mgr = wb.create_sheet("Manager Matchups")
    ws_mgr['A1'] = "MANAGER VS MANAGER MATCHUP COUNT"
    ws_mgr['A1'].font = Font(bold=True, size=12)

    row = 3
    ws_mgr[f'A{row}'] = "Manager 1"
    ws_mgr[f'B{row}'] = "Manager 2"
    ws_mgr[f'C{row}'] = "Matchups"
    for col in ['A', 'B', 'C']:
        ws_mgr[f'{col}{row}'].fill = header_fill
        ws_mgr[f'{col}{row}'].font = header_font

    row += 1
    for (m1, m2), count in manager_matchups.items():
        ws_mgr[f'A{row}'] = m1
        ws_mgr[f'B{row}'] = m2
        ws_mgr[f'C{row}'] = count
        row += 1

    ws_mgr.column_dimensions['A'].width = 20
    ws_mgr.column_dimensions['B'].width = 20
    ws_mgr.column_dimensions['C'].width = 15

    # ===== REPEATED MATCHUPS SHEET =====
    ws_repeat = wb.create_sheet("Repeated Matchups")
    ws_repeat['A1'] = "REPEATED WARRIOR MATCHUPS"
    ws_repeat['A1'].font = Font(bold=True, size=12)

    row = 3
    ws_repeat[f'A{row}'] = "Warrior 1"
    ws_repeat[f'B{row}'] = "Warrior 2"
    ws_repeat[f'C{row}'] = "Times Matched"
    for col in ['A', 'B', 'C']:
        ws_repeat[f'{col}{row}'].fill = header_fill
        ws_repeat[f'{col}{row}'].font = header_font

    row += 1
    for (w1, w2), count in repeated_matchups.items():
        ws_repeat[f'A{row}'] = w1
        ws_repeat[f'B{row}'] = w2
        ws_repeat[f'C{row}'] = count
        row += 1

    ws_repeat.column_dimensions['A'].width = 25
    ws_repeat.column_dimensions['B'].width = 25
    ws_repeat.column_dimensions['C'].width = 15

    # ===== DETAILED MATCHUPS SHEET (tier-banded, then Peasant/Monster) =====
    # Standard/challenge/blood_challenge fights grouped into same-tier and
    # adjacent cross-tier sections (Recruits, Recruits/Rookies, Rookies, ...),
    # followed by flat Peasant/Monster sections (no meaningful competitive tier,
    # so no Gap column - see build_flat_section()).
    ws_detail = wb.create_sheet("All Matchups", 0)
    headers = ["#", "Type", "Player Manager", "Player Warrior", "Record", "Tier",
               "Opponent Manager", "Opponent Warrior", "Opp Record", "Opp Tier", "Gap"]
    n_cols = len(headers)

    section_header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    section_header_font = Font(bold=True, color="FFFFFF", size=12)

    row = 1
    for section in all_sections:
        has_gap = "avg_gap" in section

        # Section title row (merged across all columns)
        ws_detail.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        title_cell = ws_detail.cell(row=row, column=1)
        count_text = f"{len(section['fights'])} fights, avg gap {section['avg_gap']:.1f}" if has_gap \
            else f"{len(section['fights'])} fights"
        title_cell.value = f"{section['label']} ({count_text})"
        title_cell.fill = section_header_fill
        title_cell.font = section_header_font
        row += 1

        # Column header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row += 1

        for idx, fight in enumerate(section["fights"], 1):
            p_tier = get_warrior_tier(fight.player_warrior)
            o_tier = get_warrior_tier(fight.opponent)
            p_record = calculate_warrior_record(fight.player_warrior)
            o_record = calculate_warrior_record(fight.opponent)

            cells_data = [
                idx,
                fight.fight_type.upper(),
                fight.player_team.manager_name,
                fight.player_warrior.name,
                p_record,
                p_tier,
                fight.opponent_manager,
                fight.opponent.name,
                o_record,
                o_tier,
                _experience_gap(fight) if has_gap else "-",
            ]

            for col_idx, value in enumerate(cells_data, 1):
                cell = ws_detail.cell(row=row, column=col_idx)
                cell.value = value
                cell.border = border

                # Color tier column
                if col_idx in (6, 10) and value in tier_colors:
                    cell.fill = PatternFill(start_color=tier_colors[value],
                                          end_color=tier_colors[value], fill_type="solid")
                    if value in ["RECRUITS", "ROOKIES"]:
                        cell.font = Font(color="000000")
                    else:
                        cell.font = Font(color="FFFFFF")

                # Color the Gap column by the same status thresholds as the web meter
                if col_idx == 11 and has_gap:
                    gap_color = _gap_status(float(value))["color"].lstrip("#")
                    cell.fill = PatternFill(start_color=gap_color, end_color=gap_color, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

            row += 1

        row += 1  # Blank spacer row between sections

    # Auto-adjust column widths
    column_widths = [5, 15, 20, 25, 12, 12, 20, 25, 12, 12, 8]
    for idx, width in enumerate(column_widths, 1):
        ws_detail.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # ===== EXPERIENCE GAP SUMMARY SHEET (table view - the accessibility twin
    # of the web UI's meter chart, per the same section breakdown) =====
    ws_gap = wb.create_sheet("Gap Summary")
    ws_gap['A1'] = "EXPERIENCE GAP BY TIER"
    ws_gap['A1'].font = Font(bold=True, size=14)
    ws_gap['A2'] = ("Average difference in total fights fought between matched warriors, per tier section. "
                     "Blood Challenge fights are excluded (no tier/experience constraint - see All Matchups sheet).")

    row = 4
    gap_headers = ["Tier Section", "Fights", "Avg Gap", "Max Gap", "Status"]
    for col_idx, header in enumerate(gap_headers, 1):
        cell = ws_gap.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    row += 1

    for section in all_sections:
        if "avg_gap" not in section:
            continue  # Peasant/Monster flat sections have no meaningful gap stat
        status = section["gap_status"]
        values = [section["label"], len(section["fights"]), round(section["avg_gap"], 2),
                  section["max_gap"], status["label"]]
        for col_idx, value in enumerate(values, 1):
            cell = ws_gap.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx == 5:
                fill_color = status["color"].lstrip("#")
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        row += 1

    ws_gap.column_dimensions['A'].width = 32
    ws_gap.column_dimensions['B'].width = 10
    ws_gap.column_dimensions['C'].width = 10
    ws_gap.column_dimensions['D'].width = 10
    ws_gap.column_dimensions['E'].width = 12

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# HTML Template
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Bloodspire Matchup Simulator</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: Arial, sans-serif; background: #f5f5f5; }
        .container { max-width: 1400px; margin: 0 auto; padding: 20px; }
        header { background: #366092; color: white; padding: 20px; margin-bottom: 20px; border-radius: 5px; }
        h1 { font-size: 28px; margin-bottom: 10px; }
        .controls { background: white; padding: 20px; margin-bottom: 20px; border-radius: 5px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        button { background: #366092; color: white; border: none; padding: 10px 20px; border-radius: 4px; cursor: pointer; font-size: 14px; margin-right: 10px; }
        button:hover { background: #2a4d6f; }
        button:disabled { background: #ccc; cursor: not-allowed; }
        label { font-weight: bold; margin-right: 10px; }
        select { padding: 8px 12px; border: 1px solid #ccc; border-radius: 4px; font-size: 14px; margin-right: 20px; }
        .status { padding: 10px; margin-bottom: 20px; border-radius: 4px; display: none; }
        .status.success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
        .status.error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
        .status.loading { background: #fff3cd; color: #856404; border: 1px solid #ffeaa7; }
        .results { background: white; padding: 20px; border-radius: 5px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); display: none; }
        .stats { display: grid; grid-template-columns: repeat(5, 1fr); gap: 15px; margin-bottom: 30px; }
        .stat-card { background: #f8f9fa; padding: 15px; border-radius: 4px; text-align: center; border-left: 4px solid #366092; }
        .stat-card h3 { color: #666; font-size: 12px; margin-bottom: 10px; text-transform: uppercase; }
        .stat-card .number { font-size: 28px; font-weight: bold; color: #366092; }
        .section { margin-bottom: 30px; }
        .section h2 { font-size: 18px; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #366092; }
        table { width: 100%; border-collapse: collapse; margin-top: 10px; }
        th { background: #366092; color: white; padding: 12px; text-align: left; font-weight: bold; }
        td { padding: 10px 12px; border-bottom: 1px solid #ddd; }
        tr:hover { background: #f5f5f5; }
        .tier { font-weight: bold; padding: 4px 8px; border-radius: 3px; color: white; }
        .tier-CHAMPION { background: #FF0000; }
        .tier-ELITES { background: #FF6600; }
        .tier-EXPERTS { background: #FFCC00; color: #000; }
        .tier-VETERANS { background: #00CC00; color: #000; }
        .tier-ADEPTS { background: #0066FF; }
        .tier-INITIATES { background: #6600FF; }
        .tier-ROOKIES { background: #999999; }
        .tier-RECRUITS { background: #CCCCCC; color: #000; }
        .fight-type { font-size: 12px; padding: 4px 8px; border-radius: 3px; font-weight: bold; }
        .fight-type-standard { background: #d4edda; color: #155724; }
        .fight-type-blood_challenge { background: #f8d7da; color: #721c24; }
        .fight-type-challenge { background: #cfe2ff; color: #084298; }
        .fight-type-monster { background: #fff3cd; color: #664d03; }
        .fight-type-peasant { background: #e2e3e5; color: #383d41; }
        .scroll-table { max-height: 600px; overflow-y: auto; margin-top: 15px; }
        .tier-sections { max-height: 1000px; overflow-y: auto; margin-top: 15px; }
        .tier-section { margin-bottom: 25px; }
        .tier-section h3 { font-size: 15px; background: #1F3864; color: white; padding: 8px 12px; border-radius: 4px 4px 0 0; margin: 0; }
        .tier-section h3 .tier-count { font-weight: normal; opacity: 0.8; font-size: 13px; }
        .tier-section-flat h3 { background: #52514e; }
        .tier-section table { margin-top: 0; }
        .tier-section-empty { color: #888; padding: 15px; text-align: center; font-style: italic; }
        /* Experience Gap meter (see: dataviz skill - meter/progress track component) */
        .gap-meter-list { margin-top: 10px; }
        .gap-meter-row { display: grid; grid-template-columns: 220px 1fr 50px 90px; align-items: center; gap: 12px; padding: 5px 0; }
        .gap-meter-label { font-size: 13px; color: #333; font-weight: 500; }
        .gap-meter-track { background: #e1e0d9; border-radius: 12px; height: 16px; overflow: hidden; }
        .gap-meter-fill { height: 100%; border-radius: 12px; }
        .gap-meter-value { font-size: 13px; color: #52514e; text-align: right; font-variant-numeric: tabular-nums; }
        .gap-meter-status { font-size: 12px; font-weight: bold; text-align: right; white-space: nowrap; }
        .gap-status-good { color: #0ca30c; }
        .gap-status-warning { color: #b3790a; }
        .gap-status-serious { color: #b8532e; }
        .gap-status-critical { color: #d03b3b; }
        .gap-badge { font-size: 11px; font-weight: bold; padding: 3px 7px; border-radius: 3px; color: #fff; }
        .loading { display: none; text-align: center; padding: 20px; }
        .spinner { border: 4px solid #f3f3f3; border-top: 4px solid #366092; border-radius: 50%; width: 40px; height: 40px; animation: spin 1s linear infinite; margin: 0 auto; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
    </style>
</head>
<body>
    <header>
        <h1>🎯 Bloodspire Matchup Simulator</h1>
        <p>Turn 12 Upload Simulation</p>
    </header>

    <div class="container">
        <div class="controls">
            <label for="turnSelect">Select Turn:</label>
            <select id="turnSelect" onchange="updateTurnDisplay()">
                <option value="">-- Select a Turn --</option>
            </select>
            <button onclick="runSimulation()">Run Simulation</button>
            <button id="exportBtn" onclick="exportToExcel()" disabled>Export to Excel</button>
            <div id="status" class="status"></div>
        </div>

        <div id="loading" class="loading">
            <div class="spinner"></div>
            <p>Running simulation...</p>
        </div>

        <div id="results" class="results">
            <div style="margin-bottom: 20px; padding: 15px; background: #e7f3ff; border-left: 4px solid #2196F3; border-radius: 4px;">
                <h3 id="turnDisplay" style="margin: 0; color: #1976D2; font-size: 18px;"></h3>
            </div>
            <div class="stats" id="statsContainer"></div>

            <div class="section">
                <h2>📊 Manager vs Manager Matchups</h2>
                <div class="scroll-table">
                    <table id="managerMatchups">
                        <thead>
                            <tr>
                                <th>Manager 1</th>
                                <th>Manager 2</th>
                                <th>Matchups</th>
                            </tr>
                        </thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>

            <div class="section">
                <h2>⚔️ Repeated Warrior Matchups</h2>
                <div class="scroll-table">
                    <table id="repeatedMatchups">
                        <thead>
                            <tr>
                                <th>Warrior 1</th>
                                <th>Warrior 2</th>
                                <th>Times Matched</th>
                            </tr>
                        </thead>
                        <tbody></tbody>
                    </table>
                </div>
            </div>

            <div class="section">
                <h2>📏 Experience Gap by Tier</h2>
                <p style="color:#666; font-size:13px; margin-bottom:10px;">
                    Average difference in total fights fought between matched warriors, per tier section below.
                    Lower is better - it means warriors are being matched with peers of similar experience,
                    not just similar tier. Blood Challenge fights are excluded from this average - they're
                    vengeance-driven with no tier/experience constraint, so they'd skew the number unfairly.
                </p>
                <div id="gapMeters" class="gap-meter-list"></div>
            </div>

            <div class="section">
                <h2>⚡ Matchups by Tier</h2>
                <p style="color:#666; font-size:13px; margin-bottom:10px;">
                    Standard, Challenge, and Blood Challenge fights, ordered lowest tier to highest, with
                    cross-tier matchups shown between each same-tier section. The Gap column shows each
                    pair's fight-count difference, but Blood Challenge fights aren't counted in a section's
                    average gap (see note above). Peasant and Monster fights don't have a meaningful
                    competitive tier, so they're listed separately below (gray headers) instead of being
                    banded into the tier ladder.
                </p>
                <div id="tieredMatchups" class="tier-sections"></div>
            </div>
        </div>
    </div>

    <script>
        // Initialize on page load
        window.addEventListener('DOMContentLoaded', async function() {
            // Load available turns
            try {
                const response = await fetch('/api/turns');
                const data = await response.json();
                const select = document.getElementById('turnSelect');

                data.turns.forEach(turn => {
                    const option = document.createElement('option');
                    option.value = turn;
                    option.textContent = `Turn ${turn}`;
                    select.appendChild(option);
                });

                // Set default to turn 12
                select.value = '12';
                updateTurnDisplay();
            } catch (error) {
                console.error('Failed to load turns:', error);
            }
        });

        function updateTurnDisplay() {
            const turn = document.getElementById('turnSelect').value;
            if (turn) {
                document.getElementById('exportBtn').disabled = true;
            }
        }

        async function runSimulation() {
            const turn = document.getElementById('turnSelect').value;
            if (!turn) {
                showStatus('error', 'Please select a turn first');
                return;
            }
            const statusEl = document.getElementById('status');
            const loadingEl = document.getElementById('loading');
            const resultsEl = document.getElementById('results');

            statusEl.style.display = 'none';
            loadingEl.style.display = 'block';
            resultsEl.style.display = 'none';

            try {
                const response = await fetch(`/api/simulate?turn=${turn}`, { method: 'POST' });
                const data = await response.json();

                if (data.error) {
                    showStatus('error', data.error);
                } else {
                    populateResults(data);
                    resultsEl.style.display = 'block';
                    document.getElementById('exportBtn').disabled = false;
                    showStatus('success', 'Simulation completed successfully!');
                }
            } catch (error) {
                showStatus('error', 'Failed to run simulation: ' + error.message);
            }

            loadingEl.style.display = 'none';
        }

        function populateResults(data) {
            // Display turn number
            document.getElementById('turnDisplay').textContent = `📋 Turn ${data.turn} Matchup Simulation`;

            // Stats
            const statsHtml = Object.entries(data.fight_types).map(([type, count]) => `
                <div class="stat-card">
                    <h3>${type}</h3>
                    <div class="number">${count}</div>
                </div>
            `).join('');
            document.getElementById('statsContainer').innerHTML = statsHtml;

            // Manager matchups
            const mgr = document.getElementById('managerMatchups').querySelector('tbody');
            mgr.innerHTML = Object.entries(data.manager_matchups).map(([pair, count]) => {
                const [m1, m2] = pair.split('|');
                return `<tr><td>${m1}</td><td>${m2}</td><td>${count}</td></tr>`;
            }).join('');

            // Repeated matchups
            const repeat = document.getElementById('repeatedMatchups').querySelector('tbody');
            repeat.innerHTML = Object.entries(data.repeated_matchups).map(([pair, count]) => {
                const [w1, w2] = pair.split('|');
                return `<tr><td>${w1}</td><td>${w2}</td><td>${count}</td></tr>`;
            }).join('');

            // Experience Gap by Tier (meter view). Only sections with gap stats
            // attached (standard/challenge fights present) are shown - a section
            // made up entirely of blood challenges has no meaningful average.
            const gapContainer = document.getElementById('gapMeters');
            const gapSections = (data.tiered_matchups || []).filter(s => s.avg_gap !== undefined && s.avg_gap !== null);
            if (!gapSections.length) {
                gapContainer.innerHTML = `<div class="tier-section-empty">No standard/challenge matchups with gap data this turn.</div>`;
            } else {
                // Scale bars against the largest gap seen (floor of 8 so a turn where
                // everything is well-matched doesn't render every bar maxed-out).
                const maxGapForScale = Math.max(8, ...gapSections.map(s => s.avg_gap));
                gapContainer.innerHTML = gapSections.map(section => {
                    const pct = Math.min(100, (section.avg_gap / maxGapForScale) * 100);
                    return `
                        <div class="gap-meter-row">
                            <div class="gap-meter-label">${section.label}</div>
                            <div class="gap-meter-track">
                                <div class="gap-meter-fill" style="width:${pct}%; background:${section.gap_status.color};"></div>
                            </div>
                            <div class="gap-meter-value">${section.avg_gap.toFixed(1)}</div>
                            <div class="gap-meter-status gap-status-${section.gap_status.key}">${section.gap_status.label}</div>
                        </div>`;
                }).join('');
            }

            // Matchups by tier, followed by Peasant/Monster fights (not tier-banded,
            // shown in their own section so they aren't dropped from the report)
            const tieredContainer = document.getElementById('tieredMatchups');
            const allSections = [...(data.tiered_matchups || []), ...(data.flat_matchups || [])];
            if (!allSections.length) {
                tieredContainer.innerHTML = `<div class="tier-section-empty">No matchups this turn.</div>`;
            } else {
                const tieredHtml = (data.tiered_matchups || []).map(s => renderMatchupSection(s, false)).join('');
                const flatHtml = (data.flat_matchups || []).map(s => renderMatchupSection(s, true)).join('');
                tieredContainer.innerHTML = tieredHtml + flatHtml;
            }
        }

        function renderMatchupSection(section, isFlat) {
            // hasGap is data-driven, not tied to isFlat: a tiered section can also
            // lack gap stats if every fight in it is a blood challenge (excluded
            // from gap analysis - see GAP_ANALYSIS_FIGHT_TYPES on the server).
            // isFlat only controls the header color (gray = not tier-banded at all).
            const hasGap = section.avg_gap !== undefined && section.avg_gap !== null;
            const rows = section.fights.map((fight, idx) => `
                <tr>
                    <td>${idx + 1}</td>
                    <td><span class="fight-type fight-type-${fight.type}">${fight.type}</span></td>
                    <td>${fight.p_manager}</td>
                    <td>${fight.p_warrior}</td>
                    <td>${fight.p_record}</td>
                    <td><span class="tier tier-${fight.p_tier}">${fight.p_tier}</span></td>
                    <td>${fight.o_manager}</td>
                    <td>${fight.o_warrior}</td>
                    <td>${fight.o_record}</td>
                    <td><span class="tier tier-${fight.o_tier}">${fight.o_tier}</span></td>
                    ${hasGap ? `<td><span class="gap-badge" style="background:${gapColor(fight.gap)};">${fight.gap}</span></td>` : ''}
                </tr>
            `).join('');
            const countText = hasGap
                ? `${section.fights.length} fights, avg gap ${section.avg_gap.toFixed(1)}`
                : `${section.fights.length} fights`;
            return `
                <div class="tier-section${isFlat ? ' tier-section-flat' : ''}">
                    <h3>${section.label} <span class="tier-count">(${countText})</span></h3>
                    <div class="scroll-table">
                        <table>
                            <thead>
                                <tr>
                                    <th>#</th><th>Type</th><th>Player Manager</th><th>Player Warrior</th>
                                    <th>Record</th><th>Tier</th><th>Opponent Manager</th>
                                    <th>Opponent Warrior</th><th>Record</th><th>Tier</th>${hasGap ? '<th>Gap</th>' : ''}
                                </tr>
                            </thead>
                            <tbody>${rows}</tbody>
                        </table>
                    </div>
                </div>`;
        }

        // Same thresholds as the server's _gap_status(): good <=1.5, warning <=3.0, serious <=5.0, else critical.
        function gapColor(gap) {
            if (gap <= 1.5) return '#0ca30c';
            if (gap <= 3.0) return '#fab219';
            if (gap <= 5.0) return '#ec835a';
            return '#d03b3b';
        }

        function showStatus(type, message) {
            const el = document.getElementById('status');
            el.className = `status ${type}`;
            el.textContent = message;
            el.style.display = 'block';
        }

        async function exportToExcel() {
            try {
                const turn = document.getElementById('turnSelect').value;
                const response = await fetch(`/api/export?turn=${turn}`);
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `turn_${turn}_matchup_simulation_${new Date().toISOString().split('T')[0]}.xlsx`;
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                a.remove();
            } catch (error) {
                showStatus('error', 'Failed to export: ' + error.message);
            }
        }
    </script>
</body>
</html>
'''

@app.before_request
def initialize_turns():
    global AVAILABLE_TURNS
    if not AVAILABLE_TURNS:
        AVAILABLE_TURNS = get_available_turns()

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/turns')
def api_get_turns():
    return jsonify({'turns': AVAILABLE_TURNS})

@app.route('/api/simulate', methods=['POST'])
def api_simulate():
    global simulation_results

    turn = request.args.get('turn', type=int, default=12)

    fights, error = run_simulation(turn)
    if error:
        return jsonify({'error': error}), 400

    organized_fights = organize_fights(fights)
    manager_matchups = calculate_manager_matchups(fights)
    repeated_matchups = find_repeated_matchups(fights)
    tiered_sections = organize_fights_by_tier(fights)
    # Peasant/Monster fights aren't part of the tier ladder (no meaningful
    # competitive tier), so they're listed as their own flat sections instead
    # of being dropped from the report entirely.
    flat_sections = [s for s in (
        build_flat_section(fights, "peasant", "Peasant Fights"),
        build_flat_section(fights, "monster", "Monster Fights"),
    ) if s]

    # Store for export (raw ScheduledFight objects - sections included so
    # the Excel export doesn't have to recompute the tier bucketing)
    simulation_results = {
        'fights': fights,
        'organized_fights': organized_fights,
        'manager_matchups': manager_matchups,
        'repeated_matchups': repeated_matchups,
        'tiered_sections': tiered_sections,
        'flat_sections': flat_sections,
    }

    def _serialize_fight(fight):
        return {
            'type': fight.fight_type,
            'p_manager': fight.player_team.manager_name,
            'p_warrior': fight.player_warrior.name,
            'p_record': calculate_warrior_record(fight.player_warrior),
            'p_fights': fight.player_warrior.total_fights,
            'p_tier': get_warrior_tier(fight.player_warrior),
            'o_manager': fight.opponent_manager,
            'o_warrior': fight.opponent.name,
            'o_record': calculate_warrior_record(fight.opponent),
            'o_fights': fight.opponent.total_fights,
            'o_tier': get_warrior_tier(fight.opponent),
            'gap': _experience_gap(fight),
        }

    def _serialize_sections(sections):
        result = []
        for section in sections:
            entry = {'label': section['label'], 'fights': [_serialize_fight(f) for f in section['fights']]}
            if 'avg_gap' in section:
                entry['avg_gap'] = round(section['avg_gap'], 2)
                entry['max_gap'] = section['max_gap']
                entry['gap_status'] = section['gap_status']
            result.append(entry)
        return result

    tiered_matchups_data = _serialize_sections(tiered_sections)
    flat_matchups_data = _serialize_sections(flat_sections)

    fight_types = {ftype: len(fights) for ftype, fights in organized_fights.items()}

    manager_pairs_formatted = {f"{m1}|{m2}": count for (m1, m2), count in manager_matchups.items()}
    repeated_pairs_formatted = {f"{w1}|{w2}": count for (w1, w2), count in repeated_matchups.items()}

    return jsonify({
        'turn': turn,
        'tiered_matchups': tiered_matchups_data,
        'flat_matchups': flat_matchups_data,
        'fight_types': fight_types,
        'total_fights': len(fights),
        'manager_matchups': manager_pairs_formatted,
        'repeated_matchups': repeated_pairs_formatted
    })

@app.route('/api/export')
def api_export():
    global simulation_results

    turn = request.args.get('turn', type=int, default=12)

    if not simulation_results:
        return jsonify({'error': 'No simulation to export'}), 400

    output = create_excel_export(
        turn,
        simulation_results['fights'],
        simulation_results['organized_fights'],
        simulation_results['manager_matchups'],
        simulation_results['repeated_matchups'],
        simulation_results['tiered_sections'] + simulation_results['flat_sections']
    )

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"matchup_simulation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

if __name__ == '__main__':
    print("=" * 80)
    print("Bloodspire Matchup Simulator - Web UI")
    print("=" * 80)
    print("\nStarting Flask server...")
    print("Open your browser to: http://localhost:5000")
    print("\nPress Ctrl+C to stop the server\n")

    app.run(debug=True, port=5000, use_reloader=False)
